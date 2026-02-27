import os
import json
from datetime import datetime
import openpyxl

RULES_JSON_PATH = os.path.join(os.path.dirname(__file__), "..", "app", "services", "classification", "classification_rules.json")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "reference", "CMA_classification.xls")

MOCK_RULES = [
    {
        "id": 1,
        "source_terms": ["Sales", "Revenue from Operations", "Turnover", "Sales A/c", "Income from Sales"],
        "target_row": 5,
        "target_sheet": "operating_statement",
        "target_label": "Net Sales / Income from Operations",
        "entity_types": ["trading", "manufacturing", "service"],
        "document_types": ["profit_and_loss"],
        "priority": 1,
        "match_type": "exact_or_fuzzy",
        "notes": "Primary revenue line. May appear as multiple variants."
    },
    {
        "id": 2,
        "source_terms": ["Purchases", "Purchase of Stock-in-Trade", "Cost of Goods Purchased"],
        "target_row": 10,
        "target_sheet": "operating_statement",
        "target_label": "Cost of Goods Sold - Purchases",
        "entity_types": ["trading"],
        "document_types": ["profit_and_loss"],
        "priority": 1,
        "match_type": "exact_or_fuzzy",
        "notes": "Trading firms only. Manufacturing firms use raw materials."
    },
    {
        "id": 3,
        "source_terms": ["Sundry Debtors", "Accounts Receivable", "Trade Receivables"],
        "target_row": 45,
        "target_sheet": "balance_sheet",
        "target_label": "Sundry Debtors",
        "entity_types": ["trading", "manufacturing", "service"],
        "document_types": ["balance_sheet", "trial_balance"],
        "priority": 1,
        "match_type": "exact_or_fuzzy",
        "notes": "Current assets"
    },
    {
        "id": 4,
        "source_terms": ["Sundry Creditors", "Accounts Payable", "Trade Payables"],
        "target_row": 115,
        "target_sheet": "balance_sheet",
        "target_label": "Sundry Creditors",
        "entity_types": ["trading", "manufacturing", "service"],
        "document_types": ["balance_sheet", "trial_balance"],
        "priority": 1,
        "match_type": "exact_or_fuzzy",
        "notes": "Current liabilities"
    },
    {
        "id": 5,
        "source_terms": ["Depreciation", "Dep.", "Depreciation A/c", "Depreciation on Fixed Assets"],
        "target_row": 25,
        "target_sheet": "operating_statement",
        "target_label": "Depreciation",
        "entity_types": ["trading", "manufacturing", "service"],
        "document_types": ["profit_and_loss"],
        "priority": 1,
        "match_type": "exact_or_fuzzy",
        "notes": "Non-cash operating expense"
    },
    {
        "id": 6,
        "source_terms": ["Salaries", "Wages", "Salary & Wages", "Employee Benefit Expenses"],
        "target_row": 18,
        "target_sheet": "operating_statement",
        "target_label": "Personnel Expenses",
        "entity_types": ["trading", "manufacturing", "service"],
        "document_types": ["profit_and_loss"],
        "priority": 2,
        "match_type": "exact_or_fuzzy",
        "notes": "Employee related costs"
    }
]

def generate_mock_rules():
    print(f"[{datetime.now()}] CMA_classification.xls not found at {EXCEL_PATH}.")
    print("Generating MOCK rules JSON instead...")
    
    output = {
        "version": "1.0",
        "total_rules": len(MOCK_RULES),
        "last_updated": datetime.now().strftime("%Y-%m-%d"),
        "rules": MOCK_RULES
    }
    
    os.makedirs(os.path.dirname(RULES_JSON_PATH), exist_ok=True)
    with open(RULES_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)
        
    print(f"\nRules converted successfully!")
    print(f"Total rules generated (MOCK): {len(MOCK_RULES)}")

def convert_excel_to_json():
    print(f"Looking for Reference CMA Rules in: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        generate_mock_rules()
        return
        
    rules = []
    # If the file did exist, we'd parse openpyxl here... Currently falling back to mock 
    # for simplicity if it fails or structure diverges.
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        
        # Read headers
        # Expecting: id, source_terms, target_row, target_sheet, target_label, entity_types, document_types, priority, match_type, notes
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            
            # Simple fallback if headers don't match the POC exactly
            # But normally we parse it into the exact JSON structure
            rule = {
                "id": int(row[0]) if row[0] else 0,
                "source_terms": [s.strip() for s in str(row[1]).split(";")] if row[1] else [],
                "target_row": int(row[2]) if row[2] else 0,
                "target_sheet": str(row[3]) if row[3] else "",
                "target_label": str(row[4]) if row[4] else "",
                "entity_types": [e.strip() for e in str(row[5]).split(",")] if row[5] else ["trading", "manufacturing", "service"],
                "document_types": [d.strip() for d in str(row[6]).split(",")] if row[6] else ["profit_and_loss", "balance_sheet"],
                "priority": int(row[7]) if row[7] else 1,
                "match_type": str(row[8]) if row[8] else "exact_or_fuzzy",
                "notes": str(row[9]) if len(row) > 9 and row[9] else ""
            }
            rules.append(rule)
            
        output = {
            "version": "1.0",
            "total_rules": len(rules),
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
            "rules": rules
        }
        
        os.makedirs(os.path.dirname(RULES_JSON_PATH), exist_ok=True)
        with open(RULES_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2)
            
        print(f"\nRules converted successfully from Excel!")
        print(f"Total rules: {len(rules)}")

    except Exception as e:
        print(f"Failed to parse Excel: {e}")
        generate_mock_rules()

if __name__ == "__main__":
    convert_excel_to_json()
