import logging
import os
from typing import Dict, Any
import openpyxl

logger = logging.getLogger(__name__)


class CMAWriter:
    def __init__(self, template_path: str) -> None:
        """Load the CMA template."""
        if not os.path.exists(template_path):
            raise FileNotFoundError("CMA template not found. Set CMA_TEMPLATE_PATH env var.")

        self.template_path = template_path
        self.workbook = openpyxl.load_workbook(template_path, keep_vba=True)

    def write(self, data: Dict[str, Any], output_path: str) -> str:
        """Write classified data into CMA template and save."""
        rows_written = 0
        rows_skipped = 0

        for sheet_name, rows_data in data.items():
            if sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                for row_idx, amount in rows_data.items():
                    try:
                        row_idx_int = int(row_idx)
                        # Column E (5) = 'Estimated Year' for V1
                        cell = ws.cell(row=row_idx_int, column=5)
                        cell.value = amount
                        rows_written += 1
                    except (ValueError, TypeError):
                        rows_skipped += 1
                        logger.warning("Skipped non-integer row key '%s' in sheet '%s'", row_idx, sheet_name)
            else:
                logger.warning("Sheet '%s' not found in CMA template", sheet_name)

        logger.info("CMA write complete: %d rows written, %d skipped", rows_written, rows_skipped)
        self.workbook.save(output_path)
        return output_path

    def get_template_info(self) -> Dict[str, Any]:
        """Return template structure info (sheets, rows)."""
        return {
            "sheets": self.workbook.sheetnames,
            "total_rows": 289,
        }
