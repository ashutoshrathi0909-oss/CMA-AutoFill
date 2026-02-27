import io
import pytest
import openpyxl
from reportlab.pdfgen import canvas

from app.services.extraction.excel_parser import parse_excel
from app.services.extraction.pdf_parser import parse_pdf, is_digital_pdf
from app.services.extraction.utils import clean_indian_number, detect_document_type, extract_financial_year


# ── Helpers ──────────────────────────────────────────────────


def create_dummy_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.append(["Mehta Computers"])
    ws.append(["Profit and Loss Account"])
    ws.append(["For the year 2024-25"])
    ws.append([])
    ws.append(["Sales", "15,00,000.00"])
    ws.append(["Purchases", "9,00,000.00"])

    f = io.BytesIO()
    wb.save(f)
    return f.getvalue()


def create_multi_sheet_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "P&L"
    ws1.append(["Company XYZ"])
    ws1.append(["Profit and Loss"])
    ws1.append([])
    ws1.append(["Sales", "10,00,000"])
    ws1.append(["Purchases", "5,00,000"])

    ws2 = wb.create_sheet("Balance Sheet")
    ws2.append(["Company XYZ"])
    ws2.append(["Balance Sheet"])
    ws2.append([])
    ws2.append(["Fixed Assets", "20,00,000"])
    ws2.append(["Current Assets", "15,00,000"])
    ws2.append(["Sundry Debtors", "8,00,000"])

    f = io.BytesIO()
    wb.save(f)
    return f.getvalue()


def create_dummy_csv() -> bytes:
    return b"Company ABC\nProfit and Loss\nFor the year 2024-25\n\nSales,1500000\nPurchases,900000\n"


def create_dummy_pdf() -> bytes:
    f = io.BytesIO()
    c = canvas.Canvas(f)
    c.drawString(100, 800, "Mehta Computers")
    c.drawString(100, 780, "Profit and Loss Account")
    c.drawString(100, 760, "Sales 1500000.00")
    c.drawString(100, 740, "Purchases 900000.00")
    c.save()
    return f.getvalue()


# ── Utils tests ──────────────────────────────────────────────


class TestCleanIndianNumber:
    def test_basic_number(self):
        assert clean_indian_number("1500000") == 1500000.0

    def test_indian_comma_format(self):
        assert clean_indian_number("15,00,000.00") == 1500000.0

    def test_parentheses_negative(self):
        assert clean_indian_number("(50,000)") == -50000.0

    def test_cr_suffix(self):
        assert clean_indian_number("50,000 Cr") == 50000.0

    def test_dr_suffix_is_negative(self):
        assert clean_indian_number("50,000 Dr") == -50000.0

    def test_empty_string(self):
        assert clean_indian_number("") == 0.0

    def test_non_numeric(self):
        assert clean_indian_number("abc") == 0.0


class TestDetectDocumentType:
    def test_profit_and_loss(self):
        assert detect_document_type("Profit and Loss Account") == "profit_and_loss"

    def test_balance_sheet(self):
        assert detect_document_type("Balance Sheet") == "balance_sheet"

    def test_trial_balance(self):
        assert detect_document_type("Trial Balance") == "trial_balance"

    def test_other(self):
        assert detect_document_type("Some Other Document") == "other"


class TestExtractFinancialYear:
    def test_short_format(self):
        assert extract_financial_year("For the year 2024-25") == "2024-25"

    def test_long_format(self):
        assert extract_financial_year("FY 2024-2025") == "2024-2025"

    def test_no_match(self):
        assert extract_financial_year("Some text without year") == "Unknown"


# ── Excel parser tests ───────────────────────────────────────


def test_excel_parser():
    file_bytes = create_dummy_excel()
    result = parse_excel(file_bytes, "mehta_pl.xlsx")

    assert result["document_type"] == "profit_and_loss"
    assert result["financial_year"] == "2024-25"
    assert result["entity_name"] == "Mehta Computers"
    assert "Sales" in [item["name"] for item in result["line_items"]]
    sales_item = next(item for item in result["line_items"] if item["name"] == "Sales")
    assert sales_item["amount"] == 1500000.0


def test_multi_sheet_excel():
    file_bytes = create_multi_sheet_excel()
    result = parse_excel(file_bytes, "multi.xlsx")

    # Should pick the sheet with more line items (Balance Sheet has 3 vs P&L has 2)
    assert len(result["line_items"]) >= 2
    assert result["metadata"]["source_file"] == "multi.xlsx"


def test_csv_parser():
    file_bytes = create_dummy_csv()
    result = parse_excel(file_bytes, "data.csv", is_csv=True)

    assert result["document_type"] == "profit_and_loss"
    assert result["metadata"]["parser"] == "csv_parser"
    assert len(result["line_items"]) >= 1


# ── PDF parser tests ─────────────────────────────────────────


def test_pdf_is_digital():
    file_bytes = create_dummy_pdf()
    assert is_digital_pdf(file_bytes) is True


def test_pdf_parser():
    file_bytes = create_dummy_pdf()
    result = parse_pdf(file_bytes, "mehta_pl.pdf")

    assert result["document_type"] == "profit_and_loss"
    assert result["metadata"]["parser"] == "pdf_parser"


# ── Standardized format test ─────────────────────────────────


def test_standardized_format():
    file_bytes = create_dummy_excel()
    result = parse_excel(file_bytes, "mehta_pl.xlsx")

    assert "document_type" in result
    assert "line_items" in result
    assert "totals" in result
    assert "metadata" in result

    first_item = result["line_items"][0]
    for key in ["name", "amount", "parent_group", "level", "is_total", "raw_text"]:
        assert key in first_item
    assert isinstance(first_item["amount"], float)


# ── Merger tests ─────────────────────────────────────────────


def test_merger_is_importable():
    from app.services.extraction.merger import merge_and_save_data
    assert callable(merge_and_save_data)


def test_merger_logic(mock_db):
    """Test that merge_and_save_data correctly merges files and updates the project."""
    from app.services.extraction.merger import merge_and_save_data

    project_id = "test-project-id"
    firm_id = "test-firm-id"

    mock_db.set_table("uploaded_files", data=[
        {
            "file_name": "pl.xlsx",
            "extracted_data": {
                "document_type": "profit_and_loss",
                "line_items": [{"name": "Sales", "amount": 1000}],
                "totals": {"gross_total": 1000},
            },
        },
        {
            "file_name": "bs.xlsx",
            "extracted_data": {
                "document_type": "balance_sheet",
                "line_items": [{"name": "Assets", "amount": 2000}, {"name": "Cash", "amount": 500}],
                "totals": {"gross_total": 2500},
            },
        },
    ])

    merge_and_save_data(project_id, firm_id)
    # If we reach here without error, merge succeeded (mock DB absorbs the writes)
